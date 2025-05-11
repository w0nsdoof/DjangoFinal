import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def generate_excel_for_approved_teams(request):
    from teams.models import Team
    teams = Team.objects.filter(status="approved").prefetch_related("members", "thesis_topic", "supervisor")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Approved Teams"

    ws.append(["№", "Студент", "Тема", "Супервайзер"])

    bold_font = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    row_num = 2
    count = 1

    for team in teams:
        supervisor = team.supervisor
        thesis = team.thesis_topic
        members = list(team.members.all())

        topic_text = (
            f"Каз: {thesis.title_kz}\n"
            f"Рус: {thesis.title_ru}\n"
            f"Англ: {thesis.title}"
        )
        supervisor_text = f"{supervisor.last_name} {supervisor.first_name}, {supervisor.degree}"

        for student in members:
            ws.append([count, f"{student.last_name} {student.first_name}", "", ""])
            row_num += 1
            count += 1

        start_merge = row_num - len(members)
        end_merge = row_num - 1
        ws.merge_cells(start_row=start_merge, end_row=end_merge, start_column=3, end_column=3)
        ws.merge_cells(start_row=start_merge, end_row=end_merge, start_column=4, end_column=4)

        topic_cell = ws.cell(row=start_merge, column=3, value=topic_text)
        topic_cell.font = Font(name='Calibri')
        topic_cell.alignment = wrap

        supervisor_cell = ws.cell(row=start_merge, column=4, value=supervisor_text)
        supervisor_cell.font = Font(name='Calibri')
        supervisor_cell.alignment = wrap

    col_widths = [5, 35, 55, 40]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    filename = f"diploma_projects_.xlsx"
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
